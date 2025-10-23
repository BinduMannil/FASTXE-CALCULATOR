"""Export utilities that build a multi-tab workbook summarising costs and pricing."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - optional dependency at runtime
    Workbook = None  # type: ignore[assignment]
    get_column_letter = None  # type: ignore[assignment]

from .calculator import BreakEvenCalculator
from .models import CostType, RevenueInputs, VendorCost


@dataclass
class ExportOptions:
    """Options that control workbook generation."""

    output_path: Path
    revenue_inputs: RevenueInputs
    expected_period_label: str = "12-month outlook"


def _ensure_workbook() -> Workbook:
    if Workbook is None:
        raise RuntimeError("openpyxl is required to export Excel workbooks. Please install the optional dependency.")
    return Workbook()


def _auto_size_columns(worksheet) -> None:
    if get_column_letter is None:
        return
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(12, length + 2)


def _costs_by_type(costs: Iterable[VendorCost], *cost_types: CostType) -> List[VendorCost]:
    selected = []
    for cost in costs:
        if cost.item.cost_type in cost_types:
            selected.append(cost)
    return selected


def _write_cost_sheet(workbook, sheet_name: str, costs: Iterable[VendorCost]) -> None:
    worksheet = workbook.create_sheet(title=sheet_name)
    headers = ["Vendor", "Item", "Type", "Amount", "Min", "Max", "Unit", "Notes", "Source"]
    worksheet.append(headers)
    for cost in costs:
        worksheet.append(
            [
                cost.vendor,
                cost.item.name,
                cost.item.cost_type.value,
                float(cost.item.amount) if cost.item.amount is not None else None,
                float(cost.item.min_amount) if cost.item.min_amount is not None else None,
                float(cost.item.max_amount) if cost.item.max_amount is not None else None,
                cost.item.unit,
                cost.item.notes,
                cost.item.source,
            ]
        )
    _auto_size_columns(worksheet)


def export_to_workbook(costs: Iterable[VendorCost], options: ExportOptions) -> Path:
    """Create an Excel workbook containing cost breakdowns and break-even metrics."""

    workbook = _ensure_workbook()
    workbook.remove(workbook.active)

    cost_list = list(costs)

    fixed_costs = _costs_by_type(
        cost_list,
        CostType.ONE_TIME,
        CostType.ANNUAL,
        CostType.SUBSCRIPTION,
        CostType.OPERATIONAL,
    )
    variable_costs = _costs_by_type(cost_list, CostType.PER_CUSTOMER, CostType.PER_TRANSACTION)

    _write_cost_sheet(workbook, "Fixed Costs", fixed_costs)
    _write_cost_sheet(workbook, "Variable Costs", variable_costs)
    _write_cost_sheet(workbook, "All Costs", cost_list)

    calculator = BreakEvenCalculator([cost.item for cost in cost_list], options.revenue_inputs)
    summary = calculator.summary()
    profitability = calculator.profitability_projection()

    summary_sheet = workbook.create_sheet(title="Summary")
    break_even_customers = calculator.break_even_customers()
    break_even_transactions = calculator.break_even_transactions()
    required_customer_price = calculator.required_customer_price()
    required_transaction_price = calculator.required_transaction_price()
    summary_sheet.append(["Metric", "Value", "Notes"])
    summary_sheet.append(["Analysis Period", options.expected_period_label, None])
    summary_sheet.append(["Fixed costs", float(summary.fixed_costs), "One-time + annual + recurring operational costs"])
    summary_sheet.append(
        [
            "Variable cost per customer",
            float(summary.variable_cost_per_customer),
            "Costs that scale with the number of active customers",
        ]
    )
    summary_sheet.append(
        [
            "Variable cost per transaction",
            float(summary.variable_cost_per_transaction),
            "Costs that scale with transaction count",
        ]
    )
    summary_sheet.append([
        "Break-even customers",
        float(break_even_customers) if break_even_customers is not None else None,
        "Number of customers needed at the configured price",
    ])
    summary_sheet.append([
        "Break-even transactions",
        float(break_even_transactions) if break_even_transactions is not None else None,
        "Number of transactions needed at the configured price",
    ])
    summary_sheet.append([
        "Required price per customer",
        float(required_customer_price) if required_customer_price is not None else None,
        "Price per customer to break even at the expected volume",
    ])
    summary_sheet.append([
        "Required price per transaction",
        float(required_transaction_price) if required_transaction_price is not None else None,
        "Price per transaction to break even at the expected volume",
    ])
    summary_sheet.append([
        "Projected revenue",
        float(profitability["revenue"]),
        "Revenue based on expected volumes and prices",
    ])
    summary_sheet.append([
        "Projected variable costs",
        float(profitability["variable_costs"]),
        "Variable costs at expected volumes",
    ])
    summary_sheet.append([
        "Projected fixed costs",
        float(profitability["fixed_costs"]),
        "Fixed costs within the analysis period",
    ])
    summary_sheet.append([
        "Projected profit",
        float(profitability["profit"]),
        "Revenue minus total costs",
    ])
    _auto_size_columns(summary_sheet)

    revenue_sheet = workbook.create_sheet(title="Revenue Inputs")
    revenue_sheet.append(["Assumption", "Value", "Notes"])
    revenue_sheet.append(["Expected customers", options.revenue_inputs.expected_customers, None])
    revenue_sheet.append(["Expected transactions", options.revenue_inputs.expected_transactions, None])
    revenue_sheet.append([
        "Customer price",
        float(options.revenue_inputs.customer_price),
        "Fee charged to each customer in the analysis period",
    ])
    revenue_sheet.append([
        "Transaction price",
        float(options.revenue_inputs.transaction_price),
        "Fee charged per transaction",
    ])
    revenue_sheet.append([
        "Subscription revenue",
        float(options.revenue_inputs.subscription_revenue),
        "Flat recurring revenue (if applicable)",
    ])
    revenue_sheet.append([
        "Analysis period (months)",
        options.revenue_inputs.analysis_period_months,
        "All costs should be expressed within this window",
    ])
    _auto_size_columns(revenue_sheet)

    output_path = options.output_path
    workbook.save(output_path)
    return output_path
